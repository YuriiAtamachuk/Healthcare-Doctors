from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import Select
from time import sleep
import openpyxl


def wait_url(driver: webdriver.Chrome, url: str):
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)

def find_element(driver: webdriver.Chrome, whichBy, unique: str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique: str) -> list[WebElement]:
    while True:
        try:
            elements =driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return elements
current_date = datetime.datetime.now().date()

wb = Workbook()
ws = wb.active

ws.merge_cells('A1:I1')
ws.merge_cells('J1:N1')
ws.merge_cells('O1:U1')
ws.merge_cells('V1:AB1')
ws.merge_cells('AC1:AI1')
ws.merge_cells('AJ1:AP1')
ws.merge_cells('AQ1:AU1')
ws.merge_cells('AV1:AZ1')
ws.merge_cells('BA1:BF1')
ws.merge_cells('BG1:BI1')
ws.merge_cells('BJ1:BM1')

ws['A1'] = "Registration details"
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color='403151', end_color='403151', fill_type="solid")
ws['A1'].font = Font(color="ffffff")
ws['J1'] = "Registration Type - General"
ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
ws['J1'].fill = PatternFill(start_color='31869B', end_color='31869B', fill_type="solid")
ws['J1'].font = Font(color="ffffff")
ws['O1'] = "Registration Type - Specialist 1"
ws['O1'].alignment = Alignment(horizontal='center', vertical='center')
ws['O1'].fill = PatternFill(start_color='244062', end_color='244062', fill_type="solid")
ws['O1'].font = Font(color="ffffff")
ws['V1'] = "Registration Type - Specialist 2"
ws['V1'].alignment = Alignment(horizontal='center', vertical='center')
ws['V1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type="solid")
ws['V1'].font = Font(color="ffffff")
ws['AC1'] = "Registration Type - Specialist 3"
ws['AC1'].alignment = Alignment(horizontal='center', vertical='center')
ws['AC1'].fill = PatternFill(start_color='244062', end_color='244062', fill_type="solid")
ws['AC1'].font = Font(color="ffffff")
ws['AJ1'] = "Registration Type - Specialist 4"
ws['AJ1'] .alignment = Alignment(horizontal='center', vertical='center')
ws['AJ1'] .fill = PatternFill(start_color='366092', end_color='366092', fill_type="solid")
ws['AJ1'].font = Font(color="ffffff")
ws['AQ1'] = "Registration Type - Provisional"
ws['AQ1'].alignment = Alignment(horizontal='center', vertical='center')
ws['AQ1'].fill = PatternFill(start_color='31869B', end_color='31869B', fill_type="solid")
ws['AQ1'].font = Font(color="ffffff")
ws['AV1'] = "Registration Type - Non-practising"
ws['AV1'].alignment = Alignment(horizontal='center', vertical='center')
ws['AV1'].fill = PatternFill(start_color='974706', end_color='974706', fill_type="solid")
ws['AV1'].font = Font(color="ffffff")
ws['BA1'] = "Registration Type - Limited"
ws['BA1'].alignment = Alignment(horizontal='center', vertical='center')
ws['BA1'].fill = PatternFill(start_color='632523', end_color='632523', fill_type="solid")
ws['BA1'].font = Font(color="ffffff")
ws['BG1'] = "Personal details"
ws['BG1'].alignment = Alignment(horizontal='center', vertical='center')
ws['BG1'].fill = PatternFill(start_color='60497A', end_color='60497A', fill_type="solid")
ws['BG1'].font = Font(color="ffffff")
ws['BJ1'] = "Principal place of practice"
ws['BJ1'].alignment = Alignment(horizontal='center', vertical='center')
ws['BJ1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type="solid")
ws['BJ1'].font = Font(color="ffffff")

ws['A2'] = "AHPRA_RegNum"
ws['B2'] = "Name"
ws['C2'] = "Profession"
ws['D2'] = "Registration status"
ws['E2'] = "Conditions"
ws['F2'] = "Undertakings"
ws['G2'] = "Reprimands"
ws['H2'] = "Tribunal decision"
ws['I2'] = "Date of first Registration in profession"
ws['J2'] = "Registration Expiry Date"
ws['K2'] = "Conditions"
ws['L2'] = "Endorsements"
ws['M2'] = "Notations"
ws['N2'] = "Registration Requirements"
ws['O2'] = "Specialty"
ws['P2'] = "Specialty Fields"
ws['Q2'] = "Registration Expiry Date"
ws['R2'] = "Conditions"
ws['S2'] = "Endorsements"
ws['T2'] = "Notations"
ws['U2'] = "Registration Requirements"
ws['V2'] = "Specialty"
ws['W2'] = "Specialty Fields"
ws['X2'] = "Registration Expiry Date"
ws['Y2'] = "Conditions"
ws['Z2'] = "Endorsements"

ws['AA2'] = "Notations"
ws['AB2'] = "Registration Requirements"
ws['AC2'] = "Specialty"
ws['AD2'] = "Specialty Fields"
ws['AE2'] = "Registration Expiry Date"
ws['AF2'] = "Conditions"
ws['AG2'] = "Endorsements"
ws['AH2'] = "Notations"
ws['AI2'] = "Registration Requirements"
ws['AJ2'] = "Specialty"
ws['AK2'] = "Specialty Fields"
ws['AL2'] = "Registration Expiry Date"
ws['AM2'] = "Conditions"
ws['AN2'] = "Endorsements"
ws['AO2'] = "Notations"
ws['AP2'] = "Registration Requirements"
ws['AQ2'] = "Registration Expiry Date"
ws['AR2'] = "Conditions"
ws['AS2'] = "Endorsements"
ws['AT2'] = "Notations"
ws['AU2'] = "Registration Requirements"
ws['AV2'] = "Registration Expiry Date"
ws['AW2'] = "Conditions"
ws['AX2'] = "Endorsements"
ws['AY2'] = "Notations"
ws['AZ2'] = "Registration Requirements"

ws['BA2'] = "Registration Expiry Date"
ws['BB2'] = "Conditions"
ws['BC2'] = "Endorsements"
ws['BD2'] = "Notations"
ws['BE2'] = "Registration Requirements"
ws['BF2'] = "Registration subtype"
ws['BG2'] = "Sex"
ws['BH2'] = "Languages (in addition to English)"
ws['BI2'] = "Qualifications"
ws['BJ2'] = "Suburb"
ws['BK2'] = "State"
ws['BL2'] = "Postcode"
ws['BM2'] = "Country"
ws['BN2'] = ""

# for col in ws.iter_cols( min_row=2, max_row=1, min_col = 3, max_col = 9):
#     for cell in col:
#         cell.fill = PatternFill(start_color="B1A0C7", end_color="B1A0C7", fill_type="solid")

# for col in ws.iter_cols(min_col = 10, max_col = 14):
#     for cell in col:
#         cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")

# for col in ws.iter_cols(min_col = 15, max_col = 21):
#     for cell in col:
#         cell.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")

# for col in ws.iter_cols(min_col = 22, max_col = 28):
#     for cell in col:
#         cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

# for col in ws.iter_cols(min_col = 29, max_col = 35):
#     for cell in col:
#         cell.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")

# for col in ws.iter_cols(min_col = 36, max_col = 42):
#     for cell in col:
#         cell.fill = PatternFill(start_color="B8CCE4", end_color="95B3D7", fill_type="solid")

# for col in ws.iter_cols(min_col = 43, max_col = 47):
#     for cell in col:
#         cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")

# for col in ws.iter_cols(min_col = 48, max_col = 52):
#     for cell in col:
#         cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")

# for col in ws.iter_cols(min_col = 53, max_col = 58):
#     for cell in col:
#         cell.fill = PatternFill(start_color="E6B8B7", end_color="FCD5B4", fill_type="solid")

# for col in ws.iter_cols(min_col = 59, max_col = 61):
#     for cell in col:
#         cell.fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")

# for col in ws.iter_cols(min_col = 62, max_col = 65):
#     for cell in col:
#         cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")


wb.save(f'ahpra_radiology_{current_date}.xlsx')

driver = webdriver.Chrome()
driver.maximize_window()


url = "https://www.ahpra.gov.au/"
driver.get(url)
sleep(1)

find_element(driver, By.CSS_SELECTOR, "#head > div.desktop-menu > div.desktop-menu__primary-mega-menu-container > div.desktop-menu__primary-nav-container > div > div > button").click()
print("Look UP!")
sleep(2)

find_element(driver, By.ID, "health-profession-dropdown").click()
print("Select Options")
sleep(1)
options = find_elements(driver, By.TAG_NAME, "li")
for option in options:
    if option.text == "Medical Practitioner":
        option.click()
        sleep(0.5)
        print("Option Clicked!")
        break
   
find_element(driver, By.ID, "predictiveSearchHomeBtn").click()
sleep(2)

num = []
people = find_elements(driver, By.CLASS_NAME, "search-results-table-row")  
for person in people:
    others = person.find_elements(By.CLASS_NAME, "search-results-table-col")
    person_data = person.find_element(By.XPATH, ".//div/div[3]")
    specialities = person_data.find_elements(By.CLASS_NAME, "col-span-row")
    for speciality in specialities:
        speciality_txt = speciality.find_element(By.CLASS_NAME, "speciality").text
        words = ["Radiology", "Radiation"]
        if any(word in speciality_txt for word in words):
            print(others[0].text)
            num.append(people.index(person))
        else:
            pass
print(num)

match_num = 0
for i in num:
    driver.get(url)
    sleep(2)

    find_element(driver, By.CSS_SELECTOR, "#head > div.desktop-menu > div.desktop-menu__primary-mega-menu-container > div.desktop-menu__primary-nav-container > div > div > button").click()
    print("Look UP!")
    sleep(2)

    find_element(driver, By.ID, "health-profession-dropdown").click()
    print("Select Options")
    sleep(1)
    options = find_elements(driver, By.TAG_NAME, "li")
    for option in options:
        if option.text == "Medical Practitioner":
            option.click()
            sleep(0.5)
            print("Option Clicked!")
            break
    
    find_element(driver, By.ID, "predictiveSearchHomeBtn").click()
    sleep(2)

    people = find_elements(driver, By.CLASS_NAME, "search-results-table-row")
    select_person = people[i]
    topersondata = select_person.find_element(By.XPATH, ".//div/div[1]/div[2]/p/a")
    topersondata.click()
    print("Clicked!")
    sleep(0.5)
    match_num += 1
    workbook = openpyxl.load_workbook(f'ahpra_radiology_{current_date}.xlsx')
    sheet = workbook['Sheet']
    dr_name = find_element(driver, By.CLASS_NAME, "practitioner-name").text
    sheet[f'B{match_num+2}'] = dr_name
    print(dr_name)
    all_data = find_elements(driver, By.CLASS_NAME, "section-row")
    for data in all_data:
        data_txt = data.find_element(By.XPATH, ".//div[1]").text
        if data_txt == "Profession":
            profession = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'C{match_num+2}'] = profession
            print(profession)
        elif data_txt == "Registration number":
            ahpra_reg_num = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'A{match_num+2}'] = ahpra_reg_num
            print(ahpra_reg_num)
        elif data_txt == "Registration status":
            reg_status = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'D{match_num+2}'] = reg_status
            print(reg_status)
        elif data_txt == 'Conditions\nCondition':
            condition = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'E{match_num+2}'] = condition
            sheet[f'K{match_num +2 }'] = condition
            sheet[f'R{match_num + 2}'] = condition
            print(condition)
        elif data_txt == 'Undertakings\nUndertaking':
            undertaking = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'F{match_num+2}'] = undertaking
            print(undertaking)
        elif data_txt == "Reprimands\nReprimand":
            reprimands = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'G{match_num + 2}'] = reprimands
            print(reprimands)
        elif data_txt == "Date of first Registration in profession\nDate of first Registration":
            first_registeration = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'I{match_num + 2}'] = first_registeration
            print(first_registeration)
        elif data_txt == "Registration Expiry Date":
            expiry_date = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'J{match_num + 2}'] = expiry_date
            print(expiry_date)
        elif data_txt == "Endorsements\nEndorsement":
            endorsements = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'L{match_num + 2}'] = endorsements
            sheet[f'S{match_num + 2}'] = endorsements
            print(endorsements)
        elif data_txt == "Notations\nNotation":
            notations = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'M{match_num + 2}'] = notations
            sheet[f'T{match_num + 2}'] = notations
            print(notations)
        elif data_txt == "Registration Requirements\nRegistration Requirement":
            reg_req = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'N{match_num + 2}'] = reg_req
            sheet[f'U{match_num + 2}'] = reg_req
            print(notations)
        elif data_txt == "Specialty":
            speciality = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'O{match_num + 2}'] = speciality
            print(speciality)
        elif data_txt == "Specialty Fields":
            speciality_field = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'P{match_num + 2}'] = speciality_field
            print(speciality_field)
        elif data_txt == "Registration Expiry Date":
            reg_expiry_date = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'Q{match_num + 2}'] = reg_expiry_date
            print(reg_expiry_date)
        elif data_txt == "Sex":
            gender = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BG{match_num + 2}'] = gender
            print(gender)
        elif data_txt == "Languages (in addition to English)":
            languages = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BH{match_num + 2}'] = languages
            print(languages)
        elif data_txt == "Qualifications":
            qualification = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BI{match_num + 2}'] = qualification
            print(qualification)
        elif data_txt == "Suburb":
            suburb = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BJ{match_num +2}'] = suburb
            print(suburb)
        elif data_txt == "State":
            state = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BK{match_num + 2}'] = state
            print(state)
        elif data_txt == "Postcode":
            postcode = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BL{match_num + 2}'] = postcode
            print(postcode)
        elif data_txt == "Country":
            country = data.find_element(By.XPATH, ".//div[2]").text
            sheet[f'BM{match_num + 2}'] = country
            print(country)
        workbook.save(f'ahpra_radiology_{current_date}.xlsx')